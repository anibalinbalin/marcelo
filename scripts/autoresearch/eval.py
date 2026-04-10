#!/usr/bin/env python3
"""eval.py — FIXED. Do not edit.

Runs one extraction experiment using the config in prompt_config.py,
compares against an Excel ground truth, and appends to log.jsonl.

Usage:
    python3 eval.py                          # uses BIMBO defaults
    python3 eval.py report.pdf golden.xlsx   # custom PDF + Excel

Output (stdout): one JSON line with score, matched, total, time_s, config.
Appends the same JSON to log.jsonl.

Environment:
    OLLAMA_URL   — default http://localhost:11434/api/generate
    MAPPINGS_KEY — company key in KNOWN_MAPPINGS (default: bimbo)
"""

import sys
import os
import json
import base64
import subprocess
import tempfile
import re
import time
from pathlib import Path
from datetime import datetime, timezone

# ── Config (agent edits this file, not here) ─────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
import prompt_config as cfg

OLLAMA_URL  = os.environ.get("OLLAMA_URL", "http://localhost:11434/api/generate")
LOG_PATH    = Path(__file__).parent / "log.jsonl"

# ── Ground truth mappings ─────────────────────────────────────────────────────
# section: BIVA section code, label: Spanish label in PDF,
# row: Excel row (1-based), transform: numeric transform to apply
KNOWN_MAPPINGS = {
    "bimbo": {
        "pdf":     "../../data/ReporteTrimestral_BIMBO_2025_4_111630-260225-2bf16217_1772054651240.pdf",
        "excel":   "../../data/BIMBOA 4Q25.xlsx",
        "col":     45,   # col AS = 4Q25
        "sheet":   "PROJ",
        # Note: [210000] balance sheet excluded — BIVA filing is standalone MX entity,
        # Excel template uses consolidated figures. They will never match.
        # Note: "Ingresos financieros" (row 15) and "Gastos financieros" (row 16)
        # excluded — Excel rows 15/16 contain a net line and None respectively,
        # not matching the separate PDF line items.
        # Note: "Gastos de venta" excluded — PDF has sales-only opex (38,945),
        # Excel row 11 is total opex = sales + admin + otros (46,780). Unmatchable.
        # Note: "Utilidad neta" uses controlling-interest specific label (row 35 = 3,157).
        "mappings": [
            {"section": "[310000]", "label": "Ingresos",                                                                  "row":  5, "transform": "div1M"},
            {"section": "[310000]", "label": "Costo de ventas",                                                           "row":  7, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Utilidad bruta",                                                            "row":  9, "transform": "div1M"},
            {"section": "[310000]", "label": "Utilidad de operacion",                                                     "row": 13, "transform": "div1M"},
            {"section": "[310000]", "label": "Participacion en la utilidad de asociadas y negocios conjuntos",            "row": 29, "transform": "div1M"},
            {"section": "[310000]", "label": "Utilidad antes de impuestos",                                               "row": 30, "transform": "div1M"},
            {"section": "[310000]", "label": "Impuestos a la utilidad",                                                   "row": 31, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Utilidad atribuible a la participacion controladora",                       "row": 35, "transform": "div1M"},
            {"section": "[310000]", "label": "Participacion no controladora",                                             "row": 33, "transform": "neg_div1M"},
        ]
    },
    "kimber": {
        "pdf":   "../../public/camila/ReporteTrimestral_KIMBER_2025_4D_111323-260211-dc210583_1770875016982 (2).pdf",
        "excel": "../../public/camila/Kimber_Fundamenta 4Q25.xlsx",
        "col":   41,   # col AO = 4Q25 (Dec 2025)
        "sheet": "PROJ",
        # Note: "Participacion en la utilidad de asociadas y negocios conjuntos" excluded
        # because KIMBER has 0 for this line — trivial match, no extraction value.
        # Note: "Participacion no controladora" excluded — also 0 for KIMBER (no minorities).
        # Note: Gastos de venta and admin mapped separately (Excel rows 10/11 split them,
        # unlike BIMBO where total opex was in one row and unmatchable).
        "mappings": [
            {"section": "[310000]", "label": "Ingresos",                                            "row":  3, "transform": "div1M"},
            {"section": "[310000]", "label": "Costo de ventas",                                     "row":  5, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Utilidad bruta",                                      "row":  6, "transform": "div1M"},
            {"section": "[310000]", "label": "Gastos de venta",                                     "row": 10, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Gastos de administracion",                            "row": 11, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Otros ingresos",                                      "row": 12, "transform": "div1M"},
            {"section": "[310000]", "label": "Utilidad de operacion",                               "row": 14, "transform": "div1M"},
            {"section": "[310000]", "label": "Ingresos financieros",                                "row": 18, "transform": "div1M"},
            {"section": "[310000]", "label": "Gastos financieros",                                  "row": 19, "transform": "neg_div1M"},
            {"section": "[310000]", "label": "Utilidad antes de impuestos",                         "row": 23, "transform": "div1M"},
            {"section": "[310000]", "label": "Impuestos a la utilidad",                             "row": 24, "transform": "neg_div1M"},
            # GLM-OCR outputs "attribuible" (double t) — use substring to avoid typo sensitivity
            {"section": "[310000]", "label": "participacion controladora",                          "row": 27, "transform": "div1M"},
        ]
    },
    "banregio": {
        "pdf":       "../../public/camila/RA Banregio 4T25.pdf",
        "excel":     "../../public/camila/RA MM ok.xlsx",
        "col":       29,    # col AC = 4T25 (Dec 2025)
        "sheet":     "PROJ",
        # PDF rounds to whole millions; Excel model has model-estimated values — use ±10.
        # Note: balance sheet excluded — PDF entity differs from Excel aggregation
        #   (TOTAL ACTIVO: PDF 277,162 vs Excel 270,095).
        # Note: "Gastos de Operacion" excluded — PDF admin-only (2,031) vs Excel broader def (2,193).
        # Note: "I.s.r. y p.t.u. causados" excluded — PDF causados-only (616) vs Excel total tax (661).
        "tolerance": 10,
        # Pages are explicit (no BIVA section codes). Image-based PDF — pdfplumber gets nothing.
        # "pages" maps section key -> list of entries. Each entry is either:
        #   - an int:   render full page
        #   - a tuple:  (page_num, y_start_pct, y_end_pct) — vertical crop
        "pages": {
            "vision:36":  [36],               # IS top: Margen financiero, Ingresos, Gastos int, PDD
            "vision:36b": [(36, 0.25, 0.55)], # IS bottom crop: Resultado, Net Income
        },
        # The quarterly income statement has 5 columns: 4T24, 1T25, 2T25, 3T25, 4T25.
        # 4T25 is the last = index 4 (0-based).
        "value_col_index": 4,
        # Values are already in MXN millions — use "passthrough" transform (just round).
        "mappings": [
            # Income Statement top half (page 36 full render)
            {"section": "vision:36",  "label": "Margen financiero",                              "row":  5, "transform": "passthrough"},
            {"section": "vision:36",  "label": "Ingresos por intereses",                         "row":  6, "transform": "passthrough"},
            {"section": "vision:36",  "label": "Gastos por intereses",                           "row":  7, "transform": "negate"},
            {"section": "vision:36",  "label": "Estimacion preventiva para riesgos crediticios", "row":  9, "transform": "negate"},
            # Income Statement bottom half (page 36 cropped 25-55%)
            {"section": "vision:36b", "label": "Resultado de la operacion",                      "row": 14, "transform": "passthrough"},
            {"section": "vision:36b", "label": "Resultado neto",                                 "row": 20, "transform": "passthrough"},
        ]
    },
    "cent": {
        "excel":  "../../data/CENT 4Q25.xlsx",
        "col":    67,
        "sheet":  "PROJ",
        "use_source_excel": True,
        "source_excel_path": "../../data/Planilha Interativa 4T25.xlsx",
        "source_sections": {
            "dre": {"sheet": "DRE I IncomeStatement", "col": 4, "label_col": 3, "data_row_start": 3},
        },
        # tolerance=1 for rounding differences (R$ thousands both sides)
        "tolerance": 1,
        "value_col_index": 0,
        # PROJ uses WITH-IFRS16 figures for most operating lines; source provides ex-IFRS16.
        # Only rows where a single source cell matches the target directly are included.
        # Excluded rows (IFRS16 adjustment gap):
        #   14 Selling expenses: src -753607 vs tgt -734318 (IFRS16 lease reclassification)
        #   15 Admin expenses:   src -165732 vs tgt -200551
        #   16 D&A:              src -57251  vs tgt -105371 (IFRS16 right-of-use depreciation)
        #   19 EBIT:             src 123582  vs tgt 165304
        #   20 Financial result: src -10442  vs tgt -51539  (IFRS16 lease finance charges)
        #   22 Finance costs:    src -98700  vs tgt -139798
        #   27/28 EBT/Tax/NI: cascade from above
        # Row 32 (Net Income with IFRS16) excluded: normalize() strips parenthesized content
        #   so "Net Income (with IFRS16)" and "Net Income (ex IFRS16)" both reduce to
        #   "net income", causing label collision — ex-IFRS16 row wins the match.
        "mappings": [
            {"section": "dre", "label": "Gross revenue",                          "row":  3, "transform": "passthrough"},
            {"section": "dre", "label": "Net revenue",                             "row":  5, "transform": "passthrough"},
            {"section": "dre", "label": "Cost of sales",                           "row":  8, "transform": "passthrough"},
            {"section": "dre", "label": "Gross profit",                            "row":  9, "transform": "passthrough"},
            {"section": "dre", "label": "Other operating income, net (ex-IFRS16)", "row": 18, "transform": "passthrough"},
            {"section": "dre", "label": "Financial Income (Expenses), net",        "row": 21, "transform": "passthrough"},
        ]
    },
    "lren3": {
        "excel":  "../../data/LREN3_4Q25_truth.xlsx",
        "col":    90,   # PROJ col CL = 4Q25 (base 1Q04 + 87 quarters)
        "sheet":  "PROJ",
        "use_source_excel": True,
        "source_excel_path": "../../public/camila/Renner Planilhas e Fundamentos  (6).xlsx",
        "source_sections": {
            # IS: labels in col B (2), 4Q25 = col 149, data starts row 7
            "is": {"sheet": "Income Statement", "col": 149, "label_col": 2, "data_row_start": 7},
            # BS: labels in col C (3, EN), 4Q25 = col 88 (datetime header Dec 31 2025), data starts row 7
            "bs": {"sheet": "Balance Sheet",    "col": 88,  "label_col": 3, "data_row_start": 7},
        },
        # R$ thousands both sides — no transform needed.
        "tolerance": 1,
        "value_col_index": 0,
        # Excluded rows (label collision — first-match wins, wrong entity):
        #   119 Noncurrent: "Noncurrent" and "Non-current" both normalize to "noncurrent";
        #       assets row 20 ("Non-current", 7,992,270) appears first and wins over
        #       liabilities row 54 ("Noncurrent ", 2,303,012). Wrong entity extracted.
        #   121 Financing - Financial Products Operations: current row 41 (21,087) appears
        #       before noncurrent row 56 (358,788) and wins. Wrong section extracted.
        #   122 Financing Lease: current row 42 (740,237) wins over noncurrent row 57
        #       (1,765,254). Wrong section extracted.
        # Row 120 (Loans, Financing and Debentures): current=0 and noncurrent=0 — both
        #       zero, collision is benign; included.
        "mappings": [
            {"section": "is", "label": "Gross Operating Revenues",              "row":   3, "transform": "passthrough"},
            {"section": "is", "label": "Costs of Goods Sold",                   "row":   7, "transform": "passthrough"},
            {"section": "is", "label": "Selling",                               "row":  11, "transform": "passthrough"},
            {"section": "is", "label": "General and Administrative",            "row":  12, "transform": "passthrough"},
            {"section": "is", "label": "Depreciation and Amortization",         "row":  13, "transform": "passthrough"},
            {"section": "is", "label": "Other Operating Income",                "row":  15, "transform": "passthrough"},
            {"section": "bs", "label": "Cash & Cash Equivalents",               "row":  90, "transform": "passthrough"},
            {"section": "bs", "label": "Short-Term Investments",                "row":  91, "transform": "passthrough"},
            {"section": "bs", "label": "Trade Accounts Receivable",             "row":  92, "transform": "passthrough"},
            {"section": "bs", "label": "Inventories",                           "row":  93, "transform": "passthrough"},
            {"section": "bs", "label": "Total Assets",                          "row": 107, "transform": "passthrough"},
            {"section": "bs", "label": "Taxes obligations",                     "row": 114, "transform": "passthrough"},
            {"section": "bs", "label": "Social and labor obligations",          "row": 115, "transform": "passthrough"},
            {"section": "bs", "label": "Liabilities Under Bylaws",              "row": 116, "transform": "passthrough"},
            {"section": "bs", "label": "Loans, Financing and Debentures",       "row": 120, "transform": "passthrough"},
            {"section": "bs", "label": "Shareholder's Equity",                  "row": 126, "transform": "passthrough"},
            {"section": "bs", "label": "Liabilities and Shareholder's Equity",  "row": 128, "transform": "passthrough"},
        ]
    },
    "natura": {
        "excel":  "../../public/camila/NATURA (limpo).xlsx",
        "col":    46,
        "sheet":  "PROJ",
        "use_source_excel": True,
        "source_excel_path": "../../public/camila/Natura Planilha de Resultados.xlsx",
        "source_sections": {
            "full_model": {"sheet": "Full model", "col": 10, "label_col": 1, "data_row_start": 3},
        },
        # tolerance=2: target values come from FAT-sheet formula chain, not direct copy.
        # Rows 4 (Deductions) and 7 (COGS) differ by 2 due to formula-chain rounding.
        # Row 19 excluded: target -514 = "Other expenses revenues" (-409) +
        #   "Transformation costs" (-105). Single-label unmatchable.
        "tolerance": 2,
        "value_col_index": 0,
        "mappings": [
            {"section": "full_model", "label": "Gross revenues",         "row":  3, "transform": "passthrough"},
            {"section": "full_model", "label": "Deductions",             "row":  4, "transform": "passthrough"},
            {"section": "full_model", "label": "COGS",                   "row":  7, "transform": "passthrough"},
            {"section": "full_model", "label": "Selling expenses",       "row": 13, "transform": "passthrough"},
            {"section": "full_model", "label": "G&A expenses",           "row": 15, "transform": "passthrough"},
            {"section": "full_model", "label": "D&A",                    "row": 17, "transform": "negate"},
            {"section": "full_model", "label": "Net financials",         "row": 26, "transform": "passthrough"},
            {"section": "full_model", "label": "Interest revenues",      "row": 27, "transform": "passthrough"},
            {"section": "full_model", "label": "Interest expenses",      "row": 28, "transform": "passthrough"},
            {"section": "full_model", "label": "Lease expenses",         "row": 29, "transform": "passthrough"},
            {"section": "full_model", "label": "Tax expenses",           "row": 34, "transform": "passthrough"},
            {"section": "full_model", "label": "Discontinued operations","row": 40, "transform": "passthrough"},
            {"section": "full_model", "label": "Net income",             "row": 42, "transform": "passthrough"},
        ]
    },
    "enel": {
        "pdf":   "../../data/enel_chile/Estados-Financieros-Enel-Chile-092025 (1).pdf",
        "excel": "../../data/enel_chile/Enel Chile Fundamenta.xlsx",
        "col":   36,    # Q3 2025 (Sep 2025)
        "sheet": "Consolidado",
        # PDF values in MUS$ (thousands of USD); Excel in USD millions → div1k.
        # Signs are preserved: costs/expenses are negative in the PDF (parentheses).
        # Note: Balance sheet excluded — PATRIMONIO TOTAL in PDF (5,470.7 USD M) differs
        #   from Excel row 60 (5,099.3) because Excel stores equity ex-NCI only.
        # Note: Row 29 "Otras ganancias" excluded — Excel row 29 aggregates it with
        #   equity-method earnings (PDF 1.9 + 3.4 = 5.4 ≠ standalone 1.9).
        "use_pdfplumber": True,
        "value_col_index": 0,   # extract_pdfplumber_q3_rows returns single-element values list
        "pages": {
            "enel:is": [6],     # IS — all Q3 2025 data
        },
        "mappings": [
            {"section": "enel:is", "label": "Total de Ingresos",                                               "row": 11, "transform": "div1k"},
            {"section": "enel:is", "label": "Materias primas y consumibles utilizados",                        "row": 15, "transform": "div1k"},
            {"section": "enel:is", "label": "Margen de Contribucion",                                          "row": 19, "transform": "div1k"},
            {"section": "enel:is", "label": "Resultado de Explotacion",                                        "row": 22, "transform": "div1k"},
            {"section": "enel:is", "label": "Ingresos financieros",                                            "row": 24, "transform": "div1k"},
            {"section": "enel:is", "label": "Costos financieros",                                              "row": 25, "transform": "div1k"},
            {"section": "enel:is", "label": "Ingreso (gasto) por impuestos a las ganancias",                   "row": 31, "transform": "div1k"},
            {"section": "enel:is", "label": "GANANCIA (PERDIDA)",                                              "row": 35, "transform": "div1k"},
            {"section": "enel:is", "label": "Ganancia (perdida) atribuible a participaciones no controladoras","row": 37, "transform": "div1k"},
            {"section": "enel:is", "label": "propietarios de la controladora",                                 "row": 36, "transform": "div1k"},
        ]
    }
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize(s):
    s = s.lower()
    if cfg.NORMALIZE_ACCENTS:
        import unicodedata
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'[^a-záéíóúñü\s]' if not cfg.NORMALIZE_ACCENTS else r'[^a-z\s]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def find_section_pages(pdf_path, target_codes):
    """Return {code: [page_nums_1indexed]} including continuation pages."""
    import pdfplumber
    pdf = pdfplumber.open(str(pdf_path))
    page_sections = {}
    tagged_pages = set()

    for i, page in enumerate(pdf.pages):
        if i == 0:
            continue
        text = page.extract_text() or ""
        for code in target_codes:
            if code in text:
                page_sections.setdefault(code, []).append(i + 1)
                tagged_pages.add(i)

    # Include continuation pages
    all_codes = set(target_codes)
    for code, pages in list(page_sections.items()):
        for page_1idx in list(pages):
            p0 = page_1idx - 1
            tables = pdf.pages[p0].extract_tables()
            if not tables:
                continue
            ref_cols = max(len(t[0]) for t in tables if t)
            nxt = p0 + 1
            if nxt < len(pdf.pages) and nxt not in tagged_pages:
                nxt_text  = pdf.pages[nxt].extract_text() or ""
                nxt_tables = pdf.pages[nxt].extract_tables()
                if nxt_tables and not any(c in nxt_text for c in all_codes):
                    nxt_cols = max(len(t[0]) for t in nxt_tables if t)
                    if abs(nxt_cols - ref_cols) <= 1 and (nxt + 1) not in pages:
                        pages.append(nxt + 1)

    pdf.close()
    return page_sections


def render_page(pdf_path, page_num, dpi, crop=None):
    """Render a PDF page as PNG. crop=(y_start_pct, y_end_pct) for vertical cropping."""
    with tempfile.TemporaryDirectory() as tmpdir:
        prefix = os.path.join(tmpdir, "pg")
        subprocess.run(
            ["pdftoppm", "-png", "-r", str(dpi), "-f", str(page_num), "-l", str(page_num), str(pdf_path), prefix],
            check=True, capture_output=True
        )
        for f in os.listdir(tmpdir):
            if f.endswith(".png"):
                full_path = os.path.join(tmpdir, f)
                if crop:
                    from PIL import Image
                    img = Image.open(full_path)
                    w, h = img.size
                    y0, y1 = int(h * crop[0]), int(h * crop[1])
                    cropped = img.crop((0, y0, w, y1))
                    buf = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
                    cropped.save(buf.name)
                    with open(buf.name, "rb") as fh:
                        data = fh.read()
                    os.unlink(buf.name)
                    return data
                with open(full_path, "rb") as fh:
                    return fh.read()
    return None


def call_glmocr(image_bytes):
    import urllib.request
    img_b64 = base64.b64encode(image_bytes).decode()
    payload = json.dumps({
        "model": "glm-ocr",
        "prompt": cfg.VLM_PROMPT,
        "images": [img_b64],
        "stream": False,
        "options": {"num_ctx": cfg.NUM_CTX, "temperature": cfg.TEMPERATURE}
    }).encode()
    req = urllib.request.Request(OLLAMA_URL, data=payload, headers={"Content-Type": "application/json"})
    resp = urllib.request.urlopen(req, timeout=300)
    result = json.loads(resp.read())
    return result.get("response", "")


def parse_html_table(html):
    from html.parser import HTMLParser
    rows, current_row, in_td, current_text = [], [], False, ""

    class P(HTMLParser):
        def handle_starttag(self, tag, attrs):  # noqa: ARG002
            nonlocal in_td, current_text
            if tag in ("td", "th"): in_td, current_text = True, ""
        def handle_endtag(self, tag):
            nonlocal in_td, current_row, current_text
            if tag in ("td", "th"):
                in_td = False
                current_row.append(current_text.strip())
            elif tag == "tr":
                if current_row: rows.append(current_row[:])
                current_row.clear()
        def handle_data(self, data):
            nonlocal current_text
            if in_td: current_text += data

    P().feed(html)
    if current_row: rows.append(current_row)

    result = []
    for row in rows:
        if len(row) < 2: continue
        values = []
        for cell in row[1:]:
            cell = cell.strip().replace(",", "").replace(" ", "")
            m = re.match(r'^\(([0-9.]+)\)$', cell)
            if m:
                values.append(-float(m.group(1)))
            else:
                try: values.append(float(cell))
                except ValueError: values.append(None)
        result.append({"label": row[0], "values": values})
    return result


def parse_plain_table(text):
    """Parse GLM-OCR plain-text table output (used for image-based PDFs like BANREGIO).

    Format per line: "Label token1 token2 ... tokenN"
    Numeric tokens: 1,234  or  (1,234)  or  -
    Label: everything before the trailing numeric run.
    """
    _num_token = re.compile(r'^\(?[\d][\d,]*\)?$|^-$')

    def parse_num(tok):
        tok = tok.replace(",", "")
        m = re.match(r'^\(([0-9.]+)\)$', tok)
        if m: return -float(m.group(1))
        if tok == "-": return None
        try: return float(tok)
        except ValueError: return None

    result = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        tokens = line.split()
        if len(tokens) < 2:
            continue
        # Walk right-to-left, consume numeric tokens
        split = len(tokens)
        for i in range(len(tokens) - 1, -1, -1):
            if _num_token.match(tokens[i]):
                split = i
            else:
                break
        label_tokens = tokens[:split]
        value_tokens = tokens[split:]
        if not label_tokens or not value_tokens:
            continue
        label  = " ".join(label_tokens)
        values = [parse_num(t) for t in value_tokens]
        result.append({"label": label, "values": values})
    return result


def parse_glmocr_output(text):
    """Route to HTML or plain-text parser depending on GLM-OCR output format."""
    if "<table" in text.lower() or "<tr" in text.lower():
        return parse_html_table(text)
    return parse_plain_table(text)


def merge_split_words(words):
    """Merge adjacent text fragments split by PDF renderer (e.g. 'M' + 'argen' → 'Margen')."""
    merged = []
    i = 0
    while i < len(words):
        w = dict(words[i])
        while i + 1 < len(words):
            nxt = words[i + 1]
            if abs(nxt['top'] - w['top']) < 2 and nxt['x0'] - w['x1'] < 1:
                w['text'] = w['text'] + nxt['text']
                w['x1'] = nxt['x1']
                i += 1
            else:
                break
        merged.append(w)
        i += 1
    return merged


def extract_source_excel_rows(source_path, sheet_name, value_col, label_col=3, data_row_start=3):
    """Read rows from a source Excel (Excel-to-Excel extraction path).
    Returns [{"label": str, "values": [val]}].
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(source_path), data_only=True)
    ws = wb[sheet_name]
    result = []
    for r in range(data_row_start, ws.max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        val   = ws.cell(row=r, column=value_col).value
        if label is not None and isinstance(val, (int, float)):
            result.append({"label": str(label).strip(), "values": [val]})
    wb.close()
    return result


def extract_pdfplumber_q3_rows(pdf_path, page_num):
    """Extract Q3-column rows from an ENEL-style IFRS statement using positional word extraction.

    Column positions (x-coords) determined from the header row of the ENEL Chile filing:
      Nota ~280-315, 9M-2025 ~315-380, 9M-2024 ~380-445, Q3-2025 ~445-515, Q3-2024 ~515+

    Returns [{"label": str, "values": [q3_val]}] — values list has one element (Q3 2025).
    Numbers use European thousands separator (period): '3.372.610' = 3372610.
    Negative values are in parentheses: '(66.511)' = -66511.
    """
    import pdfplumber
    from collections import defaultdict as _dd

    Q325_X  = (445, 515)   # Q3 2025 column
    LABEL_X = (50,  275)   # row label column

    def parse_eu_num(tok):
        tok = tok.strip()
        if tok == '-': return 0.0
        m = re.match(r'^\(([0-9.]+)\)$', tok)
        if m: return -float(m.group(1).replace('.', ''))
        try: return float(tok.replace('.', ''))
        except: return None

    pdf = pdfplumber.open(str(pdf_path))
    raw = pdf.pages[page_num - 1].extract_words(x_tolerance=3, y_tolerance=3)
    pdf.close()
    words = merge_split_words(raw)

    rows = _dd(list)
    for w in words:
        rows[round(w['top'] / 6) * 6].append(w)

    result, pending_label = [], []
    for y in sorted(rows):
        rw = sorted(rows[y], key=lambda w: w['x0'])
        label_parts, q3_val = [], None
        for w in rw:
            xm = (w['x0'] + w['x1']) / 2
            if LABEL_X[0] <= xm <= LABEL_X[1]:
                label_parts.append(w['text'])
            elif Q325_X[0] <= xm <= Q325_X[1]:
                q3_val = parse_eu_num(w['text'])
        if q3_val is not None:
            label = re.sub(r'\[.*?\]', '', ' '.join(pending_label + label_parts)).strip()
            label = re.sub(r'\s+', ' ', label).strip()
            if label:
                result.append({"label": label, "values": [q3_val]})
            pending_label = []
        elif label_parts:
            pending_label.extend(label_parts)
    return result


def apply_transform(val, transform):
    if val is None: return None
    if transform == "div1M":      return round(val / 1_000_000)
    if transform == "neg_div1M":  return round(-val / 1_000_000)
    if transform == "div1k":      return round(val / 1_000)
    if transform == "neg_div1k":  return round(-val / 1_000)
    if transform == "negate":     return round(-val)
    return round(val)  # "passthrough"


def find_best_match(target, rows):
    t = normalize(target)
    best, best_score = None, -1
    for row in rows:
        r = normalize(row["label"])
        score = 100 if r == t else (70 if t in r or r in t else 0)
        if score > best_score:
            best_score, best = score, row
    return best, best_score


def read_excel_truth(excel_path, sheet, col, mappings):
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb[sheet]
    truth = {}
    for m in mappings:
        val = ws.cell(row=m["row"], column=col).value
        truth[m["row"]] = round(val) if isinstance(val, (int, float)) else None
    wb.close()
    return truth


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    base_dir = Path(__file__).parent

    # Resolve company config
    key = os.environ.get("MAPPINGS_KEY", "bimbo")
    company = KNOWN_MAPPINGS[key]
    mappings = company["mappings"]

    excel_path = Path(sys.argv[2]) if len(sys.argv) > 2 else (base_dir / company["excel"]).resolve()

    t0 = time.time()

    if company.get("use_source_excel"):
        src_path = (base_dir / company["source_excel_path"]).resolve()
        all_rows = {}
        for sec_key, sec_cfg in company["source_sections"].items():
            all_rows[sec_key] = extract_source_excel_rows(
                src_path,
                sec_cfg["sheet"],
                sec_cfg["col"],
                sec_cfg.get("label_col", 3),
                sec_cfg.get("data_row_start", 3),
            )
    else:
        pdf_path = Path(sys.argv[1]) if len(sys.argv) > 1 else (base_dir / company["pdf"]).resolve()

        # Resolve section -> page mapping. Companies with image-based PDFs supply explicit
        # page numbers in "pages"; BIVA companies use section-code scanning.
        if "pages" in company:
            section_pages = company["pages"]
        else:
            target_codes  = list({m["section"] for m in mappings})
            section_pages = find_section_pages(pdf_path, target_codes)

        all_rows = {}
        for code, pages in section_pages.items():
            all_rows[code] = []
            for entry in pages:
                if company.get("use_pdfplumber") and isinstance(entry, int):
                    all_rows[code].extend(extract_pdfplumber_q3_rows(pdf_path, entry))
                elif isinstance(entry, tuple):
                    pnum, y0, y1 = entry
                    img = render_page(pdf_path, pnum, cfg.DPI, crop=(y0, y1))
                    if img:
                        all_rows[code].extend(parse_glmocr_output(call_glmocr(img)))
                else:
                    img = render_page(pdf_path, entry, cfg.DPI)
                    if img:
                        all_rows[code].extend(parse_glmocr_output(call_glmocr(img)))

    truth = read_excel_truth(excel_path, company["sheet"], company["col"], mappings)

    # Per-company overrides for value_col_index and match tolerance.
    value_col_index = company.get("value_col_index", cfg.VALUE_COL_INDEX)
    tolerance       = company.get("tolerance", 1)

    matched = 0
    for m in mappings:
        rows = all_rows.get(m["section"], [])
        best, score = find_best_match(m["label"], rows)
        if score < cfg.MATCH_THRESHOLD:
            continue
        idx = value_col_index
        raw = best["values"][idx] if best and len(best["values"]) > idx else (
              best["values"][0]   if best and best["values"] else None)
        transformed = apply_transform(raw, m["transform"])
        excel_val   = truth.get(m["row"])
        if transformed is not None and excel_val is not None and abs(transformed - excel_val) <= tolerance:
            matched += 1

    total   = len(mappings)
    elapsed = time.time() - t0
    score   = matched / total

    result = {
        "ts":       datetime.now(timezone.utc).isoformat(),
        "company":  key,
        "score":    round(score, 4),
        "matched":  matched,
        "total":    total,
        "time_s":   round(elapsed, 1),
        "config": {
            "VLM_PROMPT":        cfg.VLM_PROMPT,
            "DPI":               cfg.DPI,
            "NUM_CTX":           cfg.NUM_CTX,
            "TEMPERATURE":       cfg.TEMPERATURE,
            "VALUE_COL_INDEX":   cfg.VALUE_COL_INDEX,
            "MATCH_THRESHOLD":   cfg.MATCH_THRESHOLD,
            "NORMALIZE_ACCENTS": cfg.NORMALIZE_ACCENTS,
        }
    }

    print(json.dumps(result))

    with open(LOG_PATH, "a") as f:
        f.write(json.dumps(result) + "\n")


if __name__ == "__main__":
    main()
