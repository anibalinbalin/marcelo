#!/usr/bin/env python3
"""Extract data from source Excel spreadsheets using openpyxl.

Handles large files that crash ExcelJS. Reads specified sheet, finds target column,
and returns label→value pairs as JSON.

Usage: python3 extract-excel.py <xlsx_path> <sheet_name> <target_header> [label_cols]
  - xlsx_path: path to the Excel file
  - sheet_name: name of the worksheet to read
  - target_header: column header to find (e.g., "4Q25", "Q4-25")
  - label_cols: comma-separated column letters for labels (default: "B,C")

Output: JSON object { "column": <col_number>, "data": { "label": value, ... } }
"""
import openpyxl
import json
import sys
from datetime import datetime


def quarter_to_date_range(header):
    """Convert quarter header like '4Q25' to (year, quarter) for date matching."""
    import re
    m = re.match(r'^(\d)[QqTt](\d{2})$', header)
    if m:
        return (2000 + int(m.group(2)), int(m.group(1)))
    m = re.match(r'^[QqTt](\d)-(\d{2})$', header)
    if m:
        return (2000 + int(m.group(2)), int(m.group(1)))
    return None


def find_column(ws, target_header, scan_rows=(1, 2, 3, 4, 5, 6)):
    """Find the column containing the target header."""
    qinfo = quarter_to_date_range(target_header)

    for row_num in scan_rows:
        for col_idx, cell in enumerate(ws.iter_cols(min_row=row_num, max_row=row_num), 1):
            val = cell[0].value
            if val is None:
                continue

            # String match
            if str(val).strip() == target_header:
                return col_idx

            # Date match for quarter headers
            if qinfo and isinstance(val, datetime):
                yr, q = qinfo
                month = val.month
                year = val.year
                quarter_of_date = (month - 1) // 3 + 1
                if year == yr and quarter_of_date == q:
                    return col_idx

    return None


def extract(xlsx_path, sheet_name, target_header, label_cols=None):
    """Extract label→value pairs from the specified sheet and column."""
    if label_cols is None:
        label_cols = [2, 3]  # B, C by default

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(json.dumps({"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}))
        sys.exit(1)

    ws = wb[sheet_name]

    # For read_only mode, we need to iterate differently
    # First pass: find column
    col_idx = None
    rows_data = list(ws.iter_rows(values_only=True))

    for scan_row in range(min(6, len(rows_data))):
        row = rows_data[scan_row]
        for ci, val in enumerate(row):
            if val is None:
                continue
            if str(val).strip() == target_header:
                col_idx = ci
                break
            # Date match
            qinfo = quarter_to_date_range(target_header)
            if qinfo and isinstance(val, datetime):
                yr, q = qinfo
                quarter_of_date = (val.month - 1) // 3 + 1
                if val.year == yr and quarter_of_date == q:
                    col_idx = ci
                    break
        if col_idx is not None:
            break

    if col_idx is None:
        print(json.dumps({"error": f"Column '{target_header}' not found in sheet '{sheet_name}'"}))
        sys.exit(1)

    # Second pass: build label→value map AND row-indexed map
    data = {}
    row_data = {}  # row_number (1-based) → value
    for row_idx, row in enumerate(rows_data):
        if col_idx >= len(row):
            continue
        val = row[col_idx]
        if val is None or not isinstance(val, (int, float)):
            continue

        row_num = row_idx + 1  # 1-based row number
        row_data[str(row_num)] = val

        # Check label columns (0-indexed)
        for lc in label_cols:
            if lc < len(row) and row[lc]:
                label = str(row[lc]).strip().lower()
                if label and label not in ('none',):
                    data[label] = val

    wb.close()

    return {"column": col_idx, "count": len(data), "data": data, "row_data": row_data}


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python3 extract-excel.py <xlsx_path> <sheet_name> <target_header> [label_cols]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    sheet_name = sys.argv[2]
    target_header = sys.argv[3]
    label_cols = [1, 2] if len(sys.argv) < 5 else [int(x) for x in sys.argv[4].split(",")]

    result = extract(xlsx_path, sheet_name, target_header, label_cols)
    print(json.dumps(result, ensure_ascii=False))
